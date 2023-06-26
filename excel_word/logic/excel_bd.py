#!"C:\Users\vsoko\excel-word-venv\Scripts\pyhon.exe"
import openpyxl
from docx import Document

def excel_to_word():
    '''главный скрипт'''
    '''чтение документа excel'''
    excel_document = openpyxl.load_workbook(filename = "C:\\Users\\vsoko\\Documents\\графики\\График 40 отдела 2023г..xlsx", data_only=True)
    '''чтение листа из документа excel'''
    sheet = excel_document['май 2023']
    '''чтение/создание словаря вида: число:[(номер бр, фамилия), (номер бр, фамилия) и т.д.]'''
    excel_dates_dict = get_dates_dict(sheet)
    '''чтение документа word'''
    word_document = Document("C:\\Users\\vsoko\\Documents\\графики\\4 ИУ апрель 2023.docx")
    '''чтение таблицы из документа word'''
    word_table = word_document.tables[0]
    '''запись данных из excel в word'''
    write_excel_to_word(excel_dates_dict, word_table)
    return None


'''функции для записи данных по боевым постам (взятых из excel) в документ word'''
def write_excel_to_word(excel_dates_dict, word_table):
    '''формирование строки с датами из документа word в формате объектов ячеек'''
    word_date_row = word_table.rows[1].cells[3:]
    for date in word_date_row:
        date_bd = date.text.replace('\n', '')       #1, 2, 3
        bp_in_date = excel_dates_dict.get(date_bd)
        if bp_in_date is None:
            continue
        else:
            write_bp_to_word(bp_in_date, date_bd, word_date_row, word_table)
    
def write_bp_to_word(bp_in_date, date_bd, word_date_row, word_table):
    for bp, name in bp_in_date:
        print(date_bd, bp, name)
        name_row_number = get_name_row(bp, word_table)

def get_name_row(bp, word_table):
    if bp == '1':
        for index, row in enumerate(word_table.rows):
            if '1-40-1' in row.cells[0].text:
                
                if '1-40-2' in word_table.rows[index + 1].cells[0].text:
                    print('1-40-2' in word_table.rows[index + 1].cells[0].text)
                    
                    return None


'''функции для считывания данных с листа excel и формирования словаря с данными по боевым расчетам'''
def get_dates_dict(sheet): 
    '''считывание диапазона чисел и задействований'''
    dates = sheet['E15':'AI15']
    dates_dict = {}
    for i in range(0, len(dates[0])):
        date = dates[0][i]
        '''создание словаря путем применения к каждому числу месяца функции которая возвращает боевые посты'''
        dates_dict[str(date.value)] = get_bp(date, sheet)
    return dates_dict

def get_bp(date, sheet):
    '''берем число месяца и двигаемся вниз по колонке, когда встечаем номер поста то 
    вычисляем его строку и берем фамилию из колонки с "фамилиями"
    формируем кортеж типа: (номер бр, фамилия). Добавляем получившийся кортеж
    в список и в конце функции возвращаем список со всеми кортежами (номер бр: фамилия) одного дня '''
    date_column = date.column
    date_row = date.row
    bp_in_date = []
    for i in range(1, 17):
        value = str(sheet.cell(row=(date_row + i), column=date_column).value)
        if value == None:
            continue
        elif value in {'1', '2', '3', '4'} or value[:2] in {'1,', '2,', '3,', '4,'}:
            #print(value, sheet.cell(row=(date_row + i), column=4).value)
            bp_in_date.append((value[0], sheet.cell(row=(date_row + i), column=4).value))
        else: continue
    return bp_in_date


def main():
    excel_to_word()


if __name__ == "__main__":
    main()
